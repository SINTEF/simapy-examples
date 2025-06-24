"""Create an excel file with a wave scatter table.

This example shows how to create a color-coded Excel scatter table visualizing the joint distribution 
of significant wave height (Hs) and peak period (Tp) from metocean data.

Features:
    Data Retrieval:
        Uses metocean_api to get wave data for a specific location and time period
    Statistical Analysis:
        Calculates the joint frequency distribution (scatter table) using metocean_stats
    Visualization:
        Creates an Excel file with color-coded cells representing occurrence probabilities

Requirements:
    - openpyxl for Excel file creation
    - pandas for data handling
    - metocean_api and metocean_stats for data retrieval and analysis

Example:
    Basic usage of this script:
    
    >>> python create_excel_scatter.py

Note:
    Requires pandas and openpyxl to be installed:
    
    >>> pip install pandas openpyxl
"""
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from metocean_api import ts
from metocean_stats.stats.general import calculate_scatter



class ScatterExcelWriter:
    """Write a scatter table to an excel file"""

    def __init__(self, scatter: pd.DataFrame, row_name: str, column_name: str):
        self.scatter = scatter
        self.row_name = row_name
        self.column_name = column_name
        self.workbook = openpyxl.Workbook()
        self.sheet: Worksheet = self.workbook.active

    def __get_color(self, occurence: int, total: int) -> str:
        """Get a color for the cell based on the occurence and the total number of occurences"""
        if occurence == 0:
            return None

        prob = occurence / total
        # scale the probability to shift it towards the red end of the spectrum to exagerate the small values
        value = min(1.0, 5 * prob)
        red = int(255 * value)
        green = int(255 * (1 - value))
        blue = 0
        chex = [f"{i:02x}" for i in [red, green, blue]]
        return "".join(chex)

    def write_occurences(self):
        """Write the occurences to the excel file"""
        upper_row = self.scatter.index
        upper_column = self.scatter.columns
        occurences = self.scatter.values
        header = (
            [f"{self.row_name}/{self.column_name}"] + upper_column.tolist() + ["Sum"]
        )
        self.sheet.append(header)
        total_sum = occurences.sum()
        for i, occ in enumerate(occurences):
            row = [upper_row[i]] + occ.tolist() + [occ.sum()]
            self.sheet.append(row)
            for j, cell in enumerate(occ):
                color = self.__get_color(cell, total_sum)
                if color:
                    self.sheet.cell(i + 2, j + 2).fill = PatternFill(
                        "solid", start_color=color
                    )

        footer = ["Sum"] + occurences.sum(axis=0).tolist() + [total_sum]
        self.sheet.append(footer)

    def append(self, row):
        """Append a row to the excel file"""
        self.sheet.append(row)

    def save(self, filename):
        """Save the excel file"""
        self.workbook.save(filename)


def write_scatter():
    """Write a scatter table to an excel file"""
    lat_pos = 64.1154
    lon_pos = 7.8877
    start_date = "2020-10-21"
    end_date = "2020-11-21"

    product = "NORA3_wave_sub"
    var1_name = "hs"
    var2_name = "tp"

    requested_values = [var1_name, var2_name]

    df_ts = ts.TimeSeries(
        lon=lon_pos,
        lat=lat_pos,
        start_time=start_date,
        end_time=end_date,
        variable=requested_values,
        product=product,
    )

    df_ts.import_data(save_csv=True, save_nc=False, use_cache=True)

    block_size = 1.0

    scatter: pd.DataFrame = calculate_scatter(df_ts.data, var1_name, block_size, var2_name, block_size)

    writer = ScatterExcelWriter(scatter, var1_name, var2_name)
    writer.write_occurences()
    writer.append([])

    output_dir = Path("./output/simamet")
    output_dir.mkdir(exist_ok=True, parents=True)

    # Save the Excel file
    path = output_dir / f"scatter_{product}-{start_date}-{end_date}.xlsx"
    writer.save(path)


if __name__ == "__main__":
    write_scatter()
